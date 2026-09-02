"""Privacy-conscious importer for TeamWork daily schedule workbooks."""

from collections import defaultdict
from datetime import date, datetime, time, timedelta
from hashlib import sha256
from math import isfinite
from os import PathLike
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ramp_optimizer.config import TeamWorkImportConfig
from ramp_optimizer.enums import IssueSeverity, OperationalRole
from ramp_optimizer.models import (
    Employee,
    EmployeeShift,
    ImportIssue,
    ScheduleImportResult,
    VacancyRecord,
)
from ramp_optimizer.validation import validate_teamwork_import_config

REQUIRED_HEADERS = frozenset({"date", "position", "employee", "start", "end"})
OPTIONAL_HEADERS = frozenset(
    {"client", "group", "notes", "break", "hours", "swapboard"}
)
KNOWN_HEADERS = REQUIRED_HEADERS | OPTIONAL_HEADERS


def import_teamwork_schedule(
    workbook_path: str | PathLike[str],
    roster: Iterable[Employee],
    config: TeamWorkImportConfig | None = None,
) -> ScheduleImportResult:
    """Import valid schedule rows while collecting structural and row issues."""

    active_config = config or TeamWorkImportConfig()
    config_issues = validate_teamwork_import_config(active_config)
    if config_issues:
        return ScheduleImportResult(
            issues=tuple(
                ImportIssue(
                    IssueSeverity.FATAL,
                    issue.code,
                    issue.message,
                    column=issue.path,
                )
                for issue in config_issues
            )
        )

    try:
        workbook = load_workbook(
            filename=Path(workbook_path), read_only=True, data_only=True
        )
    except Exception as error:  # openpyxl exposes several format-specific errors
        return ScheduleImportResult(
            issues=(
                ImportIssue(
                    IssueSeverity.FATAL,
                    "WORKBOOK_OPEN_FAILED",
                    f"Workbook could not be opened ({type(error).__name__}).",
                ),
            )
        )

    try:
        worksheet = _find_worksheet(workbook.sheetnames, active_config.worksheet_name)
        if worksheet is None:
            return ScheduleImportResult(
                issues=(
                    ImportIssue(
                        IssueSeverity.FATAL,
                        "MISSING_SCHEDULE_WORKSHEET",
                        f"Required worksheet '{active_config.worksheet_name}' was not found.",
                    ),
                )
            )

        sheet = workbook[worksheet]
        header_row, columns, header_issues = _discover_headers(
            sheet, active_config.header_scan_limit
        )
        if header_issues:
            return ScheduleImportResult(issues=header_issues)
        assert header_row is not None

        return _import_rows(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            first_source_row=header_row + 1,
            columns=columns,
            roster=tuple(roster),
            config=active_config,
            workbook_epoch=workbook.epoch,
        )
    finally:
        workbook.close()


def normalize_header(value: object) -> str | None:
    """Normalize harmless header whitespace, punctuation, and capitalization."""

    if not isinstance(value, str):
        return None
    normalized = "".join(character for character in value if character.isalnum())
    return normalized.casefold() or None


def normalize_identity_name(value: object) -> str | None:
    """Normalize a display name for matching without changing the displayed value."""

    if not isinstance(value, str):
        return None
    normalized = " ".join(value.split()).casefold()
    return normalized or None


def normalize_position_label(value: object) -> str | None:
    """Normalize a source position solely for mapping and vacancy detection."""

    if not isinstance(value, str):
        return None
    normalized = " ".join(value.split()).casefold()
    return normalized or None


def position_mapping(config: TeamWorkImportConfig) -> Mapping[str, OperationalRole]:
    """Return the explicit, normalized position mapping configured for import."""

    return {
        normalized: role
        for label, role in config.position_role_mappings
        if (normalized := normalize_position_label(label)) is not None
    }


def _find_worksheet(sheet_names: Sequence[str], requested: str) -> str | None:
    requested_normalized = requested.strip().casefold()
    return next(
        (name for name in sheet_names if name.strip().casefold() == requested_normalized),
        None,
    )


def _discover_headers(
    worksheet: object, scan_limit: int
) -> tuple[int | None, dict[str, int], tuple[ImportIssue, ...]]:
    best_row: int | None = None
    best_columns: dict[str, int] = {}
    duplicate_headers: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True), start=1
    ):
        columns: dict[str, int] = {}
        duplicates: set[str] = set()
        for column_index, value in enumerate(row):
            normalized = normalize_header(value)
            if normalized not in KNOWN_HEADERS:
                continue
            if normalized in columns:
                duplicates.add(normalized)
            else:
                columns[normalized] = column_index

        if len(REQUIRED_HEADERS & columns.keys()) > len(
            REQUIRED_HEADERS & best_columns.keys()
        ):
            best_row = row_number
            best_columns = columns
            duplicate_headers = duplicates
        if REQUIRED_HEADERS <= columns.keys():
            if duplicates:
                return (
                    row_number,
                    columns,
                    tuple(
                        ImportIssue(
                            IssueSeverity.FATAL,
                            "DUPLICATE_HEADER",
                            f"Header '{header}' appears more than once.",
                            source_row=row_number,
                            column=header,
                        )
                        for header in sorted(duplicates)
                    ),
                )
            return row_number, columns, ()

    missing = sorted(REQUIRED_HEADERS - best_columns.keys())
    if duplicate_headers:
        missing.append("duplicate required header")
    return (
        best_row,
        best_columns,
        (
            ImportIssue(
                IssueSeverity.FATAL,
                "MISSING_REQUIRED_HEADERS",
                f"Required schedule headers were not found: {', '.join(missing)}.",
                source_row=best_row,
            ),
        ),
    )


def _import_rows(
    rows: Iterable[tuple[object, ...]],
    *,
    first_source_row: int,
    columns: Mapping[str, int],
    roster: tuple[Employee, ...],
    config: TeamWorkImportConfig,
    workbook_epoch: datetime,
) -> ScheduleImportResult:
    issues: list[ImportIssue] = []
    shifts: list[EmployeeShift] = []
    vacancies: list[VacancyRecord] = []
    seen_rows: dict[str, int] = {}
    roster_by_name: dict[str, list[Employee]] = defaultdict(list)
    for employee in roster:
        normalized_name = normalize_identity_name(employee.name)
        if normalized_name is not None:
            roster_by_name[normalized_name].append(employee)

    role_mapping = position_mapping(config)
    vacancy_labels = {
        normalized
        for label in config.vacancy_position_placeholders
        if (normalized := normalize_position_label(label)) is not None
    }

    for source_row, row in enumerate(rows, start=first_source_row):
        values = {
            header: row[index] if index < len(row) else None
            for header, index in columns.items()
        }
        if _row_is_blank(values) or _row_is_total(values):
            continue

        source_position = _display_text(values.get("position"))
        employee_name_key = normalize_identity_name(values.get("employee"))
        is_vacancy = (
            employee_name_key is None
            or normalize_position_label(source_position) in vacancy_labels
        )

        fingerprint = _row_fingerprint(values)
        if fingerprint in seen_rows:
            action = "retained as a distinct vacancy" if is_vacancy else "ignored"
            issues.append(
                ImportIssue(
                    IssueSeverity.WARNING,
                    "DUPLICATE_SCHEDULE_ROW",
                    f"Row duplicates schedule data from row {seen_rows[fingerprint]} and was {action}.",
                    source_row=source_row,
                )
            )
            if not is_vacancy:
                continue
        else:
            seen_rows[fingerprint] = source_row

        interval = _parse_interval(
            values.get("date"),
            values.get("start"),
            values.get("end"),
            workbook_epoch,
            source_row,
            config,
            issues,
        )
        if is_vacancy:
            vacancies.append(
                VacancyRecord(
                    source_row=source_row,
                    source_position=source_position,
                    start=interval[0] if interval else None,
                    end=interval[1] if interval else None,
                )
            )
            continue

        if source_position is None:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    "MISSING_POSITION",
                    "Position value is required for an occupied schedule row.",
                    source_row=source_row,
                    column="Position",
                )
            )
            continue
        if interval is None:
            continue

        matches = roster_by_name.get(employee_name_key, [])
        if not matches:
            issues.append(
                ImportIssue(
                    IssueSeverity.WARNING,
                    "UNMATCHED_EMPLOYEE",
                    f"Employee value on row {source_row} could not be matched to the roster.",
                    source_row=source_row,
                    column="Employee",
                )
            )
            continue
        if len(matches) > 1:
            issues.append(
                ImportIssue(
                    IssueSeverity.WARNING,
                    "AMBIGUOUS_EMPLOYEE",
                    f"Employee value on row {source_row} matched multiple roster records.",
                    source_row=source_row,
                    column="Employee",
                )
            )
            continue

        role = role_mapping.get(
            normalize_position_label(source_position), OperationalRole.UNKNOWN
        )
        if role is OperationalRole.UNKNOWN:
            issues.append(
                ImportIssue(
                    IssueSeverity.WARNING,
                    "UNKNOWN_POSITION",
                    f"Position value on row {source_row} requires business review and is not ramp-eligible.",
                    source_row=source_row,
                    column="Position",
                )
            )

        swapboard = _parse_swapboard(values.get("swapboard"), source_row, issues)
        _validate_hours(
            values.get("hours"), interval, source_row, config, issues
        )
        shifts.append(
            EmployeeShift(
                employee_id=matches[0].employee_id,
                start=interval[0],
                end=interval[1],
                source_position=source_position,
                normalized_role=role,
                source_row=source_row,
                notes_present=_has_value(values.get("notes")),
                swapboard=swapboard,
            )
        )

    _report_overlapping_shifts(shifts, issues)
    return ScheduleImportResult(tuple(shifts), tuple(vacancies), tuple(issues))


def _parse_interval(
    date_value: object,
    start_value: object,
    end_value: object,
    workbook_epoch: datetime,
    source_row: int,
    config: TeamWorkImportConfig,
    issues: list[ImportIssue],
) -> tuple[datetime, datetime] | None:
    parsed_date = _parse_excel_date(date_value, workbook_epoch)
    start_time = _parse_excel_time(start_value, workbook_epoch)
    end_time = _parse_excel_time(end_value, workbook_epoch)

    for column, parsed in (
        ("Date", parsed_date),
        ("Start", start_time),
        ("End", end_time),
    ):
        if parsed is None:
            issues.append(
                ImportIssue(
                    IssueSeverity.ERROR,
                    f"INVALID_{column.upper()}_VALUE",
                    f"{column} value on row {source_row} is missing or malformed.",
                    source_row=source_row,
                    column=column,
                )
            )
    if parsed_date is None or start_time is None or end_time is None:
        return None

    start = datetime.combine(parsed_date, start_time)
    end = datetime.combine(parsed_date, end_time)
    if end < start:
        end += timedelta(days=1)
    duration_hours = (end - start).total_seconds() / 3600
    if duration_hours == 0:
        issues.append(
            ImportIssue(
                IssueSeverity.ERROR,
                "ZERO_LENGTH_SHIFT",
                f"Shift on row {source_row} has zero duration and was ignored.",
                source_row=source_row,
            )
        )
        return None
    if duration_hours > config.maximum_shift_hours:
        issues.append(
            ImportIssue(
                IssueSeverity.ERROR,
                "IMPLAUSIBLY_LONG_SHIFT",
                f"Shift on row {source_row} exceeds the configured maximum duration and was ignored.",
                source_row=source_row,
            )
        )
        return None
    return start, end


def _parse_excel_date(value: object, workbook_epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if _is_number(value):
        try:
            converted = from_excel(value, workbook_epoch)
        except (TypeError, ValueError, OverflowError):
            return None
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
    return None


def _parse_excel_time(value: object, workbook_epoch: datetime) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, timedelta):
        seconds = value.total_seconds()
        if not 0 <= seconds <= 86400:
            return None
        return (datetime.min + timedelta(seconds=seconds % 86400)).time()
    if _is_number(value) and 0 <= value <= 1:
        try:
            converted = from_excel(value, workbook_epoch)
        except (TypeError, ValueError, OverflowError):
            return None
        if isinstance(converted, datetime):
            return converted.time()
        if isinstance(converted, time):
            return converted
    return None


def _parse_swapboard(
    value: object, source_row: int, issues: list[ImportIssue]
) -> bool | None:
    if not _has_value(value):
        return None
    normalized = normalize_position_label(str(value))
    if normalized in {"yes", "y", "true", "1"}:
        return True
    if normalized in {"no", "n", "false", "0"}:
        return False
    issues.append(
        ImportIssue(
            IssueSeverity.WARNING,
            "UNRECOGNIZED_SWAPBOARD_VALUE",
            f"SwapBoard value on row {source_row} was not recognized and was treated as unavailable.",
            source_row=source_row,
            column="SwapBoard",
        )
    )
    return None


def _validate_hours(
    supplied: object,
    interval: tuple[datetime, datetime],
    source_row: int,
    config: TeamWorkImportConfig,
    issues: list[ImportIssue],
) -> None:
    if not _has_value(supplied):
        return
    if not _is_number(supplied) or supplied < 0:
        issues.append(
            ImportIssue(
                IssueSeverity.WARNING,
                "INVALID_HOURS_VALUE",
                f"Hours value on row {source_row} is malformed and was not used.",
                source_row=source_row,
                column="Hours",
            )
        )
        return
    calculated = (interval[1] - interval[0]).total_seconds() / 3600
    if abs(float(supplied) - calculated) > config.hours_tolerance:
        issues.append(
            ImportIssue(
                IssueSeverity.WARNING,
                "HOURS_DURATION_MISMATCH",
                f"Hours value on row {source_row} differs from the calculated shift duration by more than {config.hours_tolerance:.2f} hours.",
                source_row=source_row,
                column="Hours",
            )
        )


def _report_overlapping_shifts(
    shifts: list[EmployeeShift], issues: list[ImportIssue]
) -> None:
    by_employee: dict[str, list[EmployeeShift]] = defaultdict(list)
    for shift in shifts:
        by_employee[shift.employee_id].append(shift)
    for employee_shifts in by_employee.values():
        ordered = sorted(employee_shifts, key=lambda shift: (shift.start, shift.end))
        for index, shift in enumerate(ordered):
            for earlier in ordered[:index]:
                if shift.start < earlier.end and earlier.start < shift.end:
                    issues.append(
                        ImportIssue(
                            IssueSeverity.ERROR,
                            "OVERLAPPING_EMPLOYEE_SHIFTS",
                            f"Shift on row {shift.source_row} overlaps another shift for the same employee on row {earlier.source_row}.",
                            source_row=shift.source_row,
                        )
                    )
                    break


def _row_is_blank(values: Mapping[str, object]) -> bool:
    return not any(_has_value(value) for value in values.values())


def _row_is_total(values: Mapping[str, object]) -> bool:
    total_labels = {"total", "totals", "grand total"}
    return any(
        normalize_position_label(values.get(column)) in total_labels
        for column in ("date", "position", "employee")
    )


def _row_fingerprint(values: Mapping[str, object]) -> str:
    serialized = "\x1f".join(
        f"{header}={_fingerprint_value(values.get(header))}"
        for header in sorted(values)
    )
    return sha256(serialized.encode("utf-8")).hexdigest()


def _fingerprint_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value.total_seconds())
    if isinstance(value, str):
        return " ".join(value.split()).casefold()
    return repr(value)


def _display_text(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    return value if value.strip() else None


def _has_value(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _is_number(value: object) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and isfinite(value)
    )
