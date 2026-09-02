"""Synthetic workbook tests for the TeamWork schedule importer."""

from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from ramp_optimizer import (
    Employee,
    IssueSeverity,
    OperationalRole,
    Qualification,
    TeamWorkImportConfig,
    import_teamwork_schedule,
)

HEADERS = [
    " Date ",
    "POSITION",
    " employee ",
    "Client",
    "Group",
    "Notes",
    "Start",
    "End",
    "Break",
    "Hours",
    "SwapBoard",
]


def roster() -> tuple[Employee, ...]:
    return (
        Employee("E001", "Avery Stone", frozenset({Qualification.PUSH})),
        Employee("E002", "Morgan Vale"),
        Employee("E003", "Rowan Hart", frozenset({Qualification.CLOSE_OUT})),
        Employee("E004", "Piper Lane"),
        Employee("E005", "Elliot Frost"),
        Employee("E006", "Jordan Twin"),
        Employee("E007", "  jordan   twin  "),
    )


def save_workbook(
    path: Path,
    rows: list[list[object]],
    *,
    headers: list[str] | None = None,
    sheet_name: str = "Schedule",
) -> Path:
    """Generate a synthetic TeamWork-shaped fixture without real employee data."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.merge_cells("A1:K1")
    worksheet["A1"] = "Synthetic TeamWork Daily Schedule"
    worksheet.merge_cells("A2:K2")
    worksheet["A2"] = "For automated tests only"
    worksheet.append([])
    worksheet.append(headers or HEADERS)
    for row in rows:
        worksheet.append(row)
    worksheet.append(["Totals"])
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def comprehensive_workbook(tmp_path: Path) -> Path:
    private_note = "PRIVATE TRAINING DETAIL MUST NOT LEAK"
    rows: list[list[object]] = [
        [date(2026, 9, 2), "Ramp Agent", "Avery Stone", None, None, private_note, time(5), time(13), None, 8, "Yes"],
        [date(2026, 9, 2), "Ramp Lead", "Avery Stone", None, None, None, time(14), time(18), None, 4, None],
        [date(2026, 9, 2), "Ramp Trainee", "Morgan Vale", None, None, None, time(16), time(0), None, 8, None],
        [date(2026, 9, 2), "Ramp Instructor", "Rowan Hart", None, None, None, time(22), time(6), None, 8, None],
        [date(2026, 9, 2), "Customer Service Agent", "Piper Lane", None, None, None, time(9), time(12), None, 3, None],
        [date(2026, 9, 2), "Unmapped Specialist", "Elliot Frost", None, None, None, time(6), time(10), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", None, None, None, None, time(5), time(9), None, 4, None],
        [date(2026, 9, 2), "(empty)", "Vacant Slot", None, None, None, time(10), time(14), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", "Avery Stone", None, None, None, time(8), time(10), None, 2, None],
        [date(2026, 9, 2), "Ramp Agent", "Avery Stone", None, None, None, time(8), time(10), None, 2, None],
        ["not-a-date", "Ramp Agent", "Morgan Vale", None, None, None, time(8), time(12), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", "Morgan Vale", None, None, None, "bad-time", time(12), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", "Unmatched Person", None, None, None, time(8), time(12), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", "Jordan Twin", None, None, None, time(8), time(12), None, 4, None],
        [date(2026, 9, 2), "Ramp Agent", "Morgan Vale", None, None, None, time(5), time(9), None, 4, "Maybe"],
        [date(2026, 9, 2), "Ramp Agent", "Morgan Vale", None, None, None, time(10), time(14), None, 3.5, None],
        [date(2026, 9, 2), "Ramp Agent", "Rowan Hart", None, None, None, time(13), time(15), None, None, None],
        [date(2026, 9, 2), "Ramp Agent", "Rowan Hart", None, None, None, time(1), time(1), None, 0, None],
        [date(2026, 9, 2), "Ramp Agent", "Rowan Hart", None, None, None, time(1), time(23), None, 22, None],
    ]
    return save_workbook(tmp_path / "synthetic-teamwork.xlsx", rows)


def issue_codes(result: object) -> set[str]:
    return {issue.code for issue in result.issues}


def test_discovers_normalized_headers_after_title_and_blank_rows(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())

    assert not result.has_fatal_issues
    assert any(shift.source_row == 5 for shift in result.shifts)


def test_missing_required_headers_is_fatal(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "missing-header.xlsx",
        [[date(2026, 9, 2), "Ramp Agent", "Avery Stone", time(5)]],
        headers=["Date", "Position", "Employee", "Start"],
    )

    result = import_teamwork_schedule(path, roster())

    assert result.has_fatal_issues
    assert issue_codes(result) == {"MISSING_REQUIRED_HEADERS"}
    assert result.shifts == ()


def test_optional_columns_may_be_absent(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "required-only.xlsx",
        [[date(2026, 9, 2), "Ramp Agent", "Avery Stone", time(5), time(13)]],
        headers=["Date", "Position", "Employee", "Start", "End"],
    )

    result = import_teamwork_schedule(path, roster())

    assert len(result.shifts) == 1
    assert result.shifts[0].notes_present is False
    assert result.shifts[0].swapboard is None


def test_excel_native_numeric_date_and_time_values_are_parsed(tmp_path: Path) -> None:
    workbook = Workbook()
    excel_date = to_excel(date(2026, 9, 2), workbook.epoch)
    workbook.close()
    path = save_workbook(
        tmp_path / "native-values.xlsx",
        [[excel_date, "Ramp Agent", "Avery Stone", 5 / 24, 13 / 24]],
        headers=["Date", "Position", "Employee", "Start", "End"],
    )

    result = import_teamwork_schedule(path, roster())

    assert len(result.shifts) == 1
    assert result.shifts[0].start == datetime(2026, 9, 2, 5)
    assert result.shifts[0].end == datetime(2026, 9, 2, 13)


def test_source_position_is_preserved_while_mapping_is_normalized(
    tmp_path: Path,
) -> None:
    path = save_workbook(
        tmp_path / "source-position.xlsx",
        [[date(2026, 9, 2), "  rAmP AgEnT  ", "Avery Stone", time(5), time(13)]],
        headers=["Date", "Position", "Employee", "Start", "End"],
    )

    result = import_teamwork_schedule(path, roster())

    assert result.shifts[0].source_position == "  rAmP AgEnT  "
    assert result.shifts[0].normalized_role is OperationalRole.RAMP_AGENT


def test_position_mapping_can_be_replaced_with_business_configuration(
    tmp_path: Path,
) -> None:
    path = save_workbook(
        tmp_path / "custom-position.xlsx",
        [[date(2026, 9, 2), "Custom Ramp", "Avery Stone", time(5), time(13)]],
        headers=["Date", "Position", "Employee", "Start", "End"],
    )
    config = TeamWorkImportConfig(
        position_role_mappings=(("Custom Ramp", OperationalRole.RAMP_AGENT),)
    )

    result = import_teamwork_schedule(path, roster(), config)

    assert result.shifts[0].normalized_role is OperationalRole.RAMP_AGENT
    assert "UNKNOWN_POSITION" not in issue_codes(result)


def test_missing_schedule_worksheet_is_fatal(tmp_path: Path) -> None:
    path = save_workbook(tmp_path / "wrong-sheet.xlsx", [], sheet_name="Other")

    result = import_teamwork_schedule(path, roster())

    assert issue_codes(result) == {"MISSING_SCHEDULE_WORKSHEET"}
    assert result.issues[0].severity is IssueSeverity.FATAL


def test_same_day_midnight_and_overnight_intervals(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    by_row = {shift.source_row: shift for shift in result.shifts}

    assert by_row[5].start == datetime(2026, 9, 2, 5)
    assert by_row[5].end == datetime(2026, 9, 2, 13)
    assert by_row[7].end == datetime(2026, 9, 3, 0)
    assert by_row[8].start == datetime(2026, 9, 2, 22)
    assert by_row[8].end == datetime(2026, 9, 3, 6)


def test_duration_blank_hours_and_invalid_shift_values_are_reported(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    codes = issue_codes(result)

    assert "HOURS_DURATION_MISMATCH" in codes
    assert "INVALID_DATE_VALUE" in codes
    assert "INVALID_START_VALUE" in codes
    assert "ZERO_LENGTH_SHIFT" in codes
    assert "IMPLAUSIBLY_LONG_SHIFT" in codes
    assert any(shift.source_row == 21 for shift in result.shifts)


def test_vacancies_are_separate_records_and_never_fictional_employees(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())

    assert [vacancy.source_row for vacancy in result.vacancies] == [11, 12]
    assert all(shift.employee_id for shift in result.shifts)
    assert not any(shift.employee_id == "Vacant Slot" for shift in result.shifts)


def test_identical_vacancy_rows_remain_distinct_records(tmp_path: Path) -> None:
    vacancy = [
        date(2026, 9, 2),
        "Ramp Agent",
        None,
        None,
        None,
        None,
        time(5),
        time(9),
        None,
        4,
        None,
    ]
    path = save_workbook(tmp_path / "vacancies.xlsx", [vacancy, vacancy.copy()])

    result = import_teamwork_schedule(path, roster())

    assert len(result.vacancies) == 2
    assert "DUPLICATE_SCHEDULE_ROW" in issue_codes(result)


def test_multiple_rows_positions_overlap_and_duplicates(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    avery_shifts = [shift for shift in result.shifts if shift.employee_id == "E001"]

    assert {shift.source_position for shift in avery_shifts} == {
        "Ramp Agent",
        "Ramp Lead",
    }
    assert len(avery_shifts) == 3
    assert "DUPLICATE_SCHEDULE_ROW" in issue_codes(result)
    assert "OVERLAPPING_EMPLOYEE_SHIFTS" in issue_codes(result)


def test_position_mapping_is_conservative_and_training_is_not_eligible_by_role(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    by_row = {shift.source_row: shift for shift in result.shifts}

    assert by_row[5].normalized_role is OperationalRole.RAMP_AGENT
    assert by_row[6].normalized_role is OperationalRole.RAMP_LEAD
    assert by_row[7].normalized_role is OperationalRole.TRAINEE
    assert by_row[8].normalized_role is OperationalRole.POSSIBLE_RAMP_SUPPORT
    assert by_row[9].normalized_role is OperationalRole.NON_RAMP
    assert by_row[10].normalized_role is OperationalRole.UNKNOWN
    assert "UNKNOWN_POSITION" in issue_codes(result)


def test_swapboard_is_parsed_conservatively(comprehensive_workbook: Path) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    by_row = {shift.source_row: shift for shift in result.shifts}

    assert by_row[5].swapboard is True
    assert by_row[19].swapboard is None
    assert "UNRECOGNIZED_SWAPBOARD_VALUE" in issue_codes(result)


def test_notes_are_reduced_to_presence_and_never_leak(
    comprehensive_workbook: Path, caplog: pytest.LogCaptureFixture
) -> None:
    sensitive = "PRIVATE TRAINING DETAIL MUST NOT LEAK"

    result = import_teamwork_schedule(comprehensive_workbook, roster())
    by_row = {shift.source_row: shift for shift in result.shifts}

    assert by_row[5].notes_present is True
    assert all(sensitive not in issue.message for issue in result.issues)
    assert sensitive not in caplog.text


def test_unmatched_and_ambiguous_names_are_private_warnings(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())
    relevant = [
        issue
        for issue in result.issues
        if issue.code in {"UNMATCHED_EMPLOYEE", "AMBIGUOUS_EMPLOYEE"}
    ]

    assert {issue.code for issue in relevant} == {
        "UNMATCHED_EMPLOYEE",
        "AMBIGUOUS_EMPLOYEE",
    }
    assert all("Unmatched Person" not in issue.message for issue in relevant)
    assert all("Jordan Twin" not in issue.message for issue in relevant)


def test_qualifications_come_only_from_roster(comprehensive_workbook: Path) -> None:
    employees = roster()
    result = import_teamwork_schedule(comprehensive_workbook, employees)

    assert all(not hasattr(shift, "qualifications") for shift in result.shifts)
    assert Qualification.PUSH in employees[0].qualifications


def test_blank_break_does_not_create_break_metadata(
    comprehensive_workbook: Path,
) -> None:
    result = import_teamwork_schedule(comprehensive_workbook, roster())

    assert all(not hasattr(shift, "break_taken") for shift in result.shifts)
